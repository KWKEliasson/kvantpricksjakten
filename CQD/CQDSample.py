import numpy as np
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import ScatterChart, Reference
# If I just import Series my IDE gives and error, but it still works
from openpyxl.chart.series_factory import SeriesFactory as Series

# Values used for xlsx formatting
ft_heading = Font(name='Calibri', bold=True)
ft_warning = Font(name='Calibri', bold=True, color='00FF0000')
wl_color = 'C0C0C0'
spec_names = {'Abs': 'Absorbans',
              'Aq': 'Vatten',
              'Cu': 'Koppar',
              'Fe': 'Järn',
              'Cd': 'Kadmium'}
chart_names = {'abs': 'Absorbans',
               'ex350': 'Fluorescens 350nm excitation',
               'ex400': 'Fluorescens 400nm excitation'}
y_names = {'abs': 'absorbans',
           'ex350': 'counts',
           'ex400': 'counts'}


class CQDSample:
    """Class representing a Carbon Quantum Dot sample"""
    def __init__(self, label, plate, well, klass, comment):
        self.label = label  # id string assigned to sample
        self.plate = plate  # index of plate where sample is found
        self.well = well  # index of well in plate where sample is found
        self.klass = klass  # school class that produced sample
        self.comment = comment  # comment entered when analysing sample
        self.spectra = {}  # dictionary of analysis spectra

    def __repr__(self):
        specs = list(self.spectra.keys())
        if len(specs) == 0:
            specs = None
        return ("<QDSample spectra:{} label:{}, klass:{}, plate:{}, well:{}, comment:{}>".
                format(specs, self.label, self.klass, self.plate, self.well, self.comment))

    def write_work_sheet(self, wb):
        """
        Append a worksheet to the workbook and write sample data to it
        :param wb: Workbook to append worksheet to
        :return: none
        """
        ws = wb.create_sheet(title=self.label)
        charts = []
        ws.column_dimensions['A'].best_fit = True
        ws.append(['Klass:', self.klass])
        ws.append(['Prov:', self.label])
        if self.comment:
            ws.append([self.comment])
        else:
            ws.append(['Inget avvikande noterat vid analys!'])
        ws.append([])  # 1 blank rows

        ws.append(['Absorbansmätning'])
        ws.cell(ws.max_row, 1).font = ft_heading
        write_spectrum(ws, charts, self, 'abs')
        ws.append([])  # 1 blank rows

        ws.append(['Flourescensmätning med 350 nm excitationsvåglängd'])
        ws.cell(ws.max_row, 1).font = ft_heading
        write_spectrum(ws, charts, self, 'ex350')
        ws.append([])  # 1 blank rows

        ws.append(['Flourescensmätning med 400 nm excitationsvåglängd'])
        ws.cell(ws.max_row, 1).font = ft_heading
        write_spectrum(ws, charts, self, 'ex400')

        anchor_row = ws.max_row+2
        anchor_cols = iter(['A', 'J', 'T'])
        for chart in charts:
            ws.add_chart(chart, '{}{}'.format(next(anchor_cols), anchor_row))


def to_list_nan(vec):  # Utility function to convert numpy arrays to lists with nan values replaced by 'Nan'
    li = vec.tolist()
    return ['Nan' if np.isnan(x) else x for x in li]


def write_spectrum(ws, charts, sample, spec_key):  # Utility function to write spectrum to worksheet
    if spec_key not in sample.spectra.keys():
        ws.append(['Ingen data'])
        ws.cell(ws.max_row, 1).font = ft_warning
        return
    spec = sample.spectra[spec_key]

    if 'gain' in spec.meta_data:
        ws.append(['Gain:', spec.meta_data['gain']])

    chart = ScatterChart()
    chart.title = chart_names[spec_key]
    chart.x_axis.title = 'våglängd [nm]'
    chart.y_axis.title = y_names[spec_key]

    ws.append(['Våglängd [nm]'] + to_list_nan(spec.wl_vector))
    for c in next(ws.iter_rows(min_row=ws.max_row, max_col=len(spec.wl_vector)+1)):
        c.fill = PatternFill('solid', fgColor=wl_color)
    xvalues = Reference(ws, min_col=2, max_col=len(spec.wl_vector)+1, min_row=ws.max_row, max_row=ws.max_row)
    chart.x_axis.scaling.min = spec.wl_vector.min()
    chart.x_axis.scaling.max = spec.wl_vector.max()

    y_vecs = spec.y_vectors
    # Background subtraction disabled
    # if spec_key == 'abs':
    #    y_vecs = spec.subtracted()

    for key, val in y_vecs.items():
        ws.append([spec_names[key]] + to_list_nan(val))
        yvalues = Reference(ws, min_col=1, max_col=len(spec.wl_vector)+1, min_row=ws.max_row, max_row=ws.max_row)
        series = Series(yvalues, xvalues, title_from_data=True)
        chart.series.append(series)

    charts.append(chart)
