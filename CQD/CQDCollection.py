import os.path
import re
import numpy as np
from openpyxl import load_workbook
from xlrd import open_workbook, XLRDError
from CQD.CQDSpectrum import CQDSpectrum
from CQD.CQDSample import CQDSample


class CQDCollection:
    """
    Class containing a collection of carbon quantum dot samples.
    It has functions to read data from a set of .xlsx and .xls files
    """

    def __init__(self):
        self.samples = []
        self.cqd1_book = None
        self.qcd2_book = None

    @classmethod
    def read_from_dir(cls, path):
        """
        Read data from a directory containing files:
            'CQD_measurements1.xls',
            'CQD_measurements2.xlsx',
            'Well plate map.xlsx',
            'Plate to Excel sheet.xlsx',
        :param path: path to directory
        :return: CQDCollection instance where .samples are populated
        """
        self = CQDCollection()
        if not os.path.isdir(path):
            raise NotADirectoryError(path + ' is not an existing directory')

        # .xls measurement data file not used anymore
        """
        cqd1_path = os.path.join(path, 'CQD_measurements1.xls')
        if not os.path.isfile(cqd1_path):
            raise FileNotFoundError(cqd1_path + ' does not exist')
        self.cqd1_book = open_workbook(cqd1_path, on_demand=True)
        """

        # cqd2_path = os.path.join(path, 'CQD_measurements2.xlsx')
        cqd2_path = os.path.join(path, 'CQD_measurements_abs adjusted_fluor adjusted.xlsx')
        if not os.path.isfile(cqd2_path):
            raise FileNotFoundError(cqd2_path + ' does not exist')
        self.qcd2_book = load_workbook(cqd2_path)

        # map_path = os.path.join(path, 'Well plate map.xlsx')
        map_path = os.path.join(path, 'Well plate map_7oct24.xlsx')
        if not os.path.isfile(map_path):
            raise FileNotFoundError(map_path + ' does not exist')
        self.parse_map_book(map_path)

        # Plate index not used anymore
        """ 
        plate_index_path = os.path.join(path, 'Plate to Excel sheet.xlsx')
        if not os.path.isfile(plate_index_path):
            raise FileNotFoundError(plate_index_path + ' does not exist')
        self.parse_plate_index(plate_index_path)
        """

        for sheet in self.qcd2_book.worksheets:
            plate = ""
            try:
                plate = int(re.match('.*?([0-9]+)$', sheet.title).group(1))
            except AttributeError:
                raise ValueError('Invalid worksheet name: {}'.format(sheet.title))
            self.parse_xlsx_sheet(sheet.title, plate)

        # cleanup to save memory
        # self.cqd1_book.release_resources()
        self.qcd2_book.close()

        flu_bg_path = os.path.join(path, 'Fluorescence baseline.xlsx')
        if not os.path.isfile(flu_bg_path):
            raise FileNotFoundError(flu_bg_path + ' does not exist')
        self.parse_flu_bg_sheet(flu_bg_path)

        # init abs background in CQDSpectrum
        self.init_abs_background()

        return self

    def parse_map_book(self, map_path):
        """
        Parses the 'Well plate map.xls' file and creates sample instances with corresponding metadata
        :param map_path: path to the map file
        """
        map_book = load_workbook(map_path)
        ms = map_book.active
        plate = None
        well_mod = None
        for row in ms.iter_rows(min_row=1, max_col=7, values_only=True):
            if row[0] == 'Plate':
                plate = row[1]  # New plate
            if row[0] == 'A':
                well_mod = 0  # First row of plate
            if row[0] == 'E':
                well_mod = 12  # Second row of plate

            reactants = []
            if row[3] is not None:
                reactants.append(row[3])
            if row[4] is not None:
                reactants.append(row[4])

            if row[2] is not None:
                klass = row[5].strip().upper()
                self.samples.append(CQDSample(str(row[2]), plate, well_mod + row[1], klass, row[6], reactants))
        map_book.close()

    def parse_plate_index(self, plate_index_path):
        """
        Parses the 'Well plate map.xlsx' file to find the sheet containing the spectrum data for each plate.
        Then calls functions to parse each sheet.
        :param plate_index_path: path to the plate index file
        """

        def get_trailing_digits(s):
            try:
                return re.match('.*?([0-9]+)$', s).group(1)
            except AttributeError:
                raise ValueError('Plate {} has invalid sheet index description'.format(plate))

        index_book = load_workbook(plate_index_path)
        ws = index_book.active
        for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
            plate = row[0]
            if plate is None:
                continue
            if row[1] is not None and row[2] is None:
                sheet_ind = get_trailing_digits(row[1])
                self.parse_xls_sheet("Sheet{}".format(sheet_ind), plate)
            elif row[2] is not None and row[1] is None:
                sheet_ind = get_trailing_digits(row[2])
                self.parse_xlsx_sheet("Sheet{}".format(sheet_ind), plate)
            else:
                raise ValueError('Plate {} has invalid sheet index description'.format(plate))
        index_book.close()

    def parse_xls_sheet(self, sheet_name, plate):
        """
        Extracts data from a xls sheet and creates CQDSpectrum instances that are added to the corresponding samples
        :param sheet_name: Name of worksheet to parse
        :param plate: Plate index corresponding to worksheet
        :return:
        """
        try:
            ws = self.cqd1_book.sheet_by_name(sheet_name)
        except XLRDError:
            print('Could not find sheet {} in QCD1 workbook'.format(sheet_name))
            return
        fc = ws.col(0)
        for i in (x for x in fc if x.value.startswith('Label: ')):
            l_row = fc.index(i)
            st_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Start Time:')))
            et_row = fc.index(next((x for x in fc[l_row:] if x.value == 'End Time:')))
            wl_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Wavel.')))

            rows = []
            while fc[wl_row].ctype != 0:
                rows.append(ws.row_values(wl_row))
                wl_row += 1

            spectrum = CQDSpectrum.spectrum_from_xl_data(ws.cell(st_row, 1).value, ws.cell(et_row, 1).value,
                                                         ws.col_values(0, l_row + 1, st_row),
                                                         ws.col_values(4, l_row + 1, st_row),
                                                         rows)
            well, spec_type = parse_spectrometer_label(i.value)
            samples = list(s for s in self.samples if s.plate == plate and s.well == well)
            if len(samples) != 1:
                raise ValueError('plate={}, well={} finds {} samples. Not exactly 1!'.format(plate, well, len(samples)))
            samples[0].spectra[spec_type] = spectrum
            self.cqd1_book.unload_sheet(sheet_name)  # close sheet to save memory

    def parse_xlsx_sheet(self, sheet_name, plate):
        """
        Extracts data from a xlsx sheet and creates CQDSpectrum instances that are added to the corresponding samples
        :param sheet_name: Name of worksheet to parse
        :param plate: Plate index corresponding to worksheet
        :return:
        """
        try:
            ws = self.qcd2_book[sheet_name]
        except KeyError:
            print('Could not find sheet {} in QCD2 workbook'.format(sheet_name))
            return
        fc = ws['A']
        for i in (x for x in fc if x.value is not None and x.value.startswith('Label: ')):

            well, spec_type = parse_spectrometer_label(i.value)
            samples = list(s for s in self.samples if s.plate == plate and s.well == well)
            if len(samples) == 0:
                print('plate={}, well={} no sample initialized. Discarding data!'.format(plate, well))
                continue
            elif len(samples) > 1:
                raise ValueError('plate={}, well={} finds {} samples. More than 1!'.format(plate, well, len(samples)))

            l_row = fc.index(i)
            st_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Start Time:')))
            et_row = fc.index(next((x for x in fc[l_row:] if x.value == 'End Time:')))
            wl_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Wavel.')))

            rows = []
            while fc[wl_row].value is not None:
                r = list(ws.iter_rows(min_row=wl_row + 1, max_row=wl_row + 1, values_only=True))[0]
                try:
                    end_i = r.index(None)
                    rows.append(list(r[:end_i]))
                except ValueError:
                    rows.append(list(r))
                wl_row += 1

            attr_col = list(ws.iter_cols(min_row=l_row + 2, max_row=st_row, min_col=1, max_col=1, values_only=True))
            val_col = list(ws.iter_cols(min_row=l_row + 2, max_row=st_row, min_col=5, max_col=5, values_only=True))

            spectrum = CQDSpectrum.spectrum_from_xl_data(ws.cell(st_row + 1, 2).value, ws.cell(et_row + 1, 2).value,
                                                         list(attr_col[0]), list(val_col[0]), rows)
            samples[0].spectra[spec_type] = spectrum

    def parse_flu_bg_sheet(self, flu_bg_path):
        wb = load_workbook(flu_bg_path)
        ws = wb.active
        fc = ws['A']
        for i in (x for x in fc if x.value is not None and x.value.startswith('Label: ')):
            l_row = fc.index(i)
            st_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Start Time:')))
            et_row = fc.index(next((x for x in fc[l_row:] if x.value == 'End Time:')))
            wl_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Wavel.')))

            rows = []
            while fc[wl_row].value is not None:
                r = list(ws.iter_rows(min_row=wl_row + 1, max_row=wl_row + 1, values_only=True))[0]
                try:
                    end_i = r.index(None)
                    rows.append(list(r[:end_i]))
                except ValueError:
                    rows.append(list(r))
                wl_row += 1

            attr_col = list(ws.iter_cols(min_row=l_row + 2, max_row=st_row, min_col=1, max_col=1, values_only=True))
            val_col = list(ws.iter_cols(min_row=l_row + 2, max_row=st_row, min_col=5, max_col=5, values_only=True))

            spectrum = CQDSpectrum.spectrum_from_xl_data(ws.cell(st_row + 1, 2).value, ws.cell(et_row + 1, 2).value,
                                                         list(attr_col[0]), list(val_col[0]),
                                                         rows)
            if spectrum.meta_data['ex_wl'] == 350:
                if CQDSpectrum.bg_ex350_x is None:
                    CQDSpectrum.bg_ex350_x = spectrum.wl_vector
                elif not np.array_equal(CQDSpectrum.bg_ex350_x, spectrum.wl_vector):
                    raise ValueError("Trying to initialize fluorescence background with invalid wavelength vector")
                CQDSpectrum.bg_ex350_ys[spectrum.meta_data['gain']] = spectrum.y_vectors
            elif spectrum.meta_data['ex_wl'] == 400:
                if CQDSpectrum.bg_ex400_x is None:
                    CQDSpectrum.bg_ex400_x = spectrum.wl_vector
                elif not np.array_equal(CQDSpectrum.bg_ex400_x, spectrum.wl_vector):
                    raise ValueError("Trying to initialize fluorescence background with invalid wavelength vector")
                CQDSpectrum.bg_ex400_ys[spectrum.meta_data['gain']] = spectrum.y_vectors
            else:
                raise ValueError('Spectrum with invalid wavelength {} in "Fluorescence baseline.xlsx" file'.
                                 format(spectrum.meta_data['ex_wl']))
        wb.close()

    def init_abs_background(self):
        """
        Initialize the absorbance background of the CQDSpectrum class from blank sample absorbance spectra
        :return: Nothing
        """
        blank_samples = []
        blank_sample_comments = ['no sample in eppendorfrör (analysis is of water only)',
                                 'sample missing (analysis is of water only)',
                                 'Eppendorfrör empty - analysis is of water only',
                                 'sample and protokoll missing (analysis is of water only)',
                                 'sample missing - analysis is of water only',
                                 'No sample, synthesis did not work - analysis is of water only',
                                 'reference sample with only water']
        _ = [blank_samples.append(x) for x in self.samples if x.comment in blank_sample_comments]
        bg_abs_ys = []
        bg_abs_x = None
        for samp in blank_samples:
            if 'abs' not in samp.spectra:
                continue
            bg_abs_ys.append(samp.spectra['abs'].y_vectors['Abs'])
            if bg_abs_x is None:
                bg_abs_x = samp.spectra['abs'].wl_vector
            elif not np.array_equal(bg_abs_x, samp.spectra['abs'].wl_vector):
                raise ValueError('Error initializing absorbance background. Sample {} has different wl_vector'
                                 .format(samp))

        bg_abs_ys = np.array(bg_abs_ys)
        mean_y = bg_abs_ys.mean(axis=0)
        CQDSpectrum.bg_abs_x = bg_abs_x
        CQDSpectrum.bg_abs_y = mean_y


def parse_spectrometer_label(label):
    """
    Utility function from reading spectrometer data labels.
    :param label: Label string from spectrometer data sheet
    :return: well: index of well, spec_type: String describing measurement type
    """
    label = label.strip()
    spec_type = 'abs'
    if label[-1].isalpha():
        well = int(label[6:-1])
        if label[-1] == 'a':
            spec_type = 'ex350'
        elif label[-1] == 'b':
            spec_type = 'ex400'
    else:
        well = int(label[6:])
    return well, spec_type
