import numpy
import numpy as np
import os.path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from xlrd import open_workbook, XLRDError
import re
import time


class CQDCollection:
    """
    Class containing a collection of carbon quantum dot samples.
    It has functions to read data from a set of .xlsx and .xls files
    """
    def __init__(self):
        self.samples = []
        qcd1_book = None
        qcd2_book = None

    @classmethod
    def read_from_dir(cls, path):
        """
        Read data from a directory containing files:
            'CQD_measurements1.xls',
            'CQD_measurements2.xlsx',
            'Well plate map.xlsx',
            'Plate to Excel sheet.xlsx',
        :param path: path to directory
        :return: CQDCollection instance where .samples are populated
        """
        self = CQDCollection()
        if not os.path.isdir(path):
            raise NotADirectoryError(path + ' is not an existing directory')
        cqd1_path = os.path.join(path, 'CQD_measurements1.xls')
        if not os.path.isfile(cqd1_path):
            raise FileNotFoundError(cqd1_path + ' does not exist')
        self.qcd1_book = open_workbook(cqd1_path)

        cqd2_path = os.path.join(path, 'CQD_measurements2.xlsx')
        if not os.path.isfile(cqd2_path):
            raise FileNotFoundError(cqd2_path + ' does not exist')
        self.qcd2_book = load_workbook(cqd2_path)

        map_path = os.path.join(path, 'Well plate map.xlsx')
        if not os.path.isfile(map_path):
            raise FileNotFoundError(map_path + ' does not exist')
        self.parse_map_book(map_path)

        plate_index_path = os.path.join(path, 'Plate to Excel sheet.xlsx')
        if not os.path.isfile(plate_index_path):
            raise FileNotFoundError(plate_index_path + ' does not exist')
        self.parse_plate_index(plate_index_path)

        # cleanup to save memory
        self.qcd1_book = None
        self.qcd2_book = None

        return self

    def parse_map_book(self, map_path):
        """
        Parses the 'Well plate map.xls' file and creates sample instances with corresponding metadata
        :param map_path: path to the map file
        """
        map_book = load_workbook(map_path)
        ms = map_book.active
        plate = None
        klass = None
        well_mod = None
        for row in ms.iter_rows(min_row=1, max_col=5, values_only=True):
            if row[0] == 'Plate':
                plate = row[1]  # New plate
            if row[0] == 'A':
                well_mod = 0  # First row of plate
            if row[0] == 'E':
                well_mod = 12  # Second row of plate

            if row[3] is not None:
                klass = row[3]  # New klass

            if row[2] is not None:
                self.samples.append(CQDSample(str(row[2]), plate, well_mod + row[1], klass, row[4]))  # append sample

    def parse_plate_index(self, plate_index_path):
        """
        Parses the 'Well plate map.xlsx' file to find the sheet containing the spectrum data for each plate.
        Then calls functions to parse each sheet.
        :param plate_index_path: path to the plate index file
        """
        def get_trailing_digits(s):
            try:
                return re.match('.*?([0-9]+)$', s).group(1)
            except AttributeError:
                raise ValueError('Plate {} has invalid sheet index description'.format(plate))

        index_book = load_workbook(plate_index_path)
        ws = index_book.active
        for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
            plate = row[0]
            if plate is None:
                continue
            if row[1] is not None and row[2] is None:
                sheet_ind = get_trailing_digits(row[1])
                self.parse_xls_sheet("Sheet{}".format(sheet_ind), plate)
            elif row[2] is not None and row[1] is None:
                sheet_ind = get_trailing_digits(row[2])
                self.parse_xlsx_sheet("Sheet{}".format(sheet_ind), plate)
            else:
                raise ValueError('Plate {} has invalid sheet index description'.format(plate))

    def parse_xls_sheet(self, sheet_name, plate):
        """
        Extracts data from a xls sheet and creates CQDSpectrum instances that are added to the corresponding samples
        :param sheet_name: Name of worksheet to parse
        :param plate: Plate index corresponding to worksheet
        :return:
        """
        try:
            ws = self.qcd1_book.sheet_by_name(sheet_name)
        except XLRDError as e:
            print('Could not find sheet {} in QCD1 workbook'.format(sheet_name))
            return
        fc = ws.col(0)
        for i in (x for x in fc if x.value.startswith('Label: ')):
            l_row = fc.index(i)
            st_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Start Time:')))
            et_row = fc.index(next((x for x in fc[l_row:] if x.value == 'End Time:')))
            wl_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Wavel.')))

            rows = []
            while fc[wl_row].ctype != 0:
                rows.append(ws.row_values(wl_row))
                wl_row += 1

            spectrum = CQDSpectrum.spectrum_from_xl_data(ws.cell(st_row, 1).value, ws.cell(et_row, 1).value,
                                  ws.col_values(0, l_row+1, st_row),
                                  ws.col_values(4, l_row+1, st_row),
                                  rows)
            well, spec_type = parse_spectrometer_label(i.value)
            samples = list(s for s in self.samples if s.plate == plate and s.well == well)
            if len(samples) != 1:
                raise ValueError('plate={}, well={} finds {} samples. Not exactly 1!'.format(plate, well, len(samples)))
            samples[0].spectra[spec_type] = spectrum

    def parse_xlsx_sheet(self, sheet_name, plate):
        """
        Extracts data from a xlsx sheet and creates CQDSpectrum instances that are added to the corresponding samples
        :param sheet_name: Name of worksheet to parse
        :param plate: Plate index corresponding to worksheet
        :return:
        """
        try:
            ws = self.qcd2_book[sheet_name]
        except KeyError as e:
            print('Could not find sheet {} in QCD2 workbook'.format(sheet_name))
            return
        fc = ws['A']
        for i in (x for x in fc if x.value is not None and x.value.startswith('Label: ')):
            l_row = fc.index(i)
            st_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Start Time:')))
            et_row = fc.index(next((x for x in fc[l_row:] if x.value == 'End Time:')))
            wl_row = fc.index(next((x for x in fc[l_row:] if x.value == 'Wavel.')))

            rows = []
            while fc[wl_row].value != None:
                r = list(ws.iter_rows(min_row=wl_row+1, max_row=wl_row+1, values_only=True))[0]
                try:
                    end_i = r.index(None)
                    rows.append(list(r[:end_i]))
                except ValueError:
                    rows.append(list(r))
                wl_row += 1

            attr_col = list(ws.iter_cols(min_row=l_row+2, max_row=st_row, min_col=1, max_col=1, values_only=True))
            val_col = list(ws.iter_cols(min_row=l_row+2, max_row=st_row, min_col=5, max_col=5, values_only=True))

            spectrum = CQDSpectrum.spectrum_from_xl_data(ws.cell(st_row+1, 2).value, ws.cell(et_row+1, 2).value,
                                             list(attr_col[0]), list(val_col[0]),
                                             rows)
            well, spec_type = parse_spectrometer_label(i.value)
            samples = list(s for s in self.samples if s.plate == plate and s.well == well)
            if len(samples) != 1:
                raise ValueError('plate={}, well={} finds {} samples. Not exactly 1!'.format(plate, well, len(samples)))
            samples[0].spectra[spec_type] = spectrum


def parse_spectrometer_label(label):
    """
    Utility function from reading spectrometer data labels.
    :param label: Label string from spectrometer data sheet
    :return: well: index of well, spec_type: String describing measurement type
    """
    label = label.strip()
    spec_type = 'abs'
    if label[-1].isalpha():
        well = int(label[6:-1])
        if label[-1] == 'a':
            spec_type = 'ex350'
        elif label[-1] == 'b':
            spec_type = 'ex400'
    else:
        well = int(label[6:])
    return well, spec_type


def row_to_np_array(row: list) -> object:
    """
    Utility function that converts a list of values from an Excel worksheet to a numpy array
    :param row: list of cell values from an Excel worksheet
    :return: numpy array
    """
    row = row[1:]
    i = 0
    for c in row:
        if c == 'OVER':
            row[i] = 'Nan'
        if c == '':
            row = row[:i]
            break
        i += 1
    return np.array(row, dtype=float)


class CQDSample:
    """Class representing a Carbon Quantum Dot sample"""
    def __init__(self, label, plate, well, klass, comment):
        self.label = label  # id string assigned to sample
        self.plate = plate  # index of plate where sample is found
        self.well = well  # index of well in plate where sample is found
        self.klass = klass  # school class that produced sample
        self.comment = comment  # comment entered when analysing sample
        self.spectra = {}  # dictionary of analysis spectra

    def __repr__(self):
        specs = list(self.spectra.keys())
        if len(specs) == 0:
            specs = None
        return "<QDSample spectra:{} label:{}, klass:{}, plate:{}, well:{}, comment:{}>".format(specs, self.label, self.klass, self.plate, self.well, self.comment)

    def write_work_sheet(self, wb):
        """
        Append a worksheet to the workbook and write sample data to it
        :param wb: Workbook to append worksheet to
        :return: none
        """

        ft_heading = Font(name='Calibri', bold=True)
        ft_warning = Font(name='Calibri', bold=True, color='00FF0000')
        wl_color = 'C0C0C0'
        spec_names = {'Abs': 'Absorbans',
                      'Aq': 'Vatten',
                      'Cu': 'Koppar',
                      'Fe': 'Järn',
                      'Cd': 'Kadmium'}
        chart_names = {'abs': 'Absorbtion',
                       'ex350': 'Fluorescens 350nm excitation',
                       'ex400': 'Fluorescens 400nm excitation'}
        y_names = {'abs': 'absorbans',
                   'ex350': 'counts',
                   'ex400': 'counts'}

        def to_list_nan(vec):  # Utility function to convert numpy arrays to lists with nan values replaced by 'Nan'
            li = vec.tolist()
            return ['Nan' if np.isnan(x) else x for x in li]

        def write_spectrum(sample, spec_key):  # Utility function to write spectrum to worksheet
            if spec_key not in self.spectra.keys():
                ws.append(['Ingen data'])
                ws.cell(ws.max_row, 1).font = ft_warning
                return
            spec = sample.spectra[spec_key]

            if 'gain' in spec.meta_data:
                ws.append(['Gain:', spec.meta_data['gain']])

            chart = ScatterChart()
            chart.title = chart_names[spec_key]
            chart.x_axis.title = 'våglängd [nm]'
            chart.y_axis.title = y_names[spec_key]

            ws.append(['Våglängd [nm]'] + to_list_nan(spec.wl_vector))
            for c in next(ws.iter_rows(min_row=ws.max_row, max_col=len(spec.wl_vector)+1)):
                c.fill = PatternFill('solid', fgColor=wl_color)
            xvalues = Reference(ws, min_col=1, max_col=len(spec.wl_vector)+1, min_row=ws.max_row, max_row=ws.max_row)

            for key, val in spec.y_vectors.items():
                ws.append([spec_names[key]] + to_list_nan(val))
                yvalues = Reference(ws, min_col=1, max_col=len(spec.wl_vector)+1, min_row=ws.max_row, max_row=ws.max_row)
                series = Series(yvalues, xvalues, title_from_data=True)
                chart.series.append(series)

            charts.append(chart)

        ws = wb.create_sheet(title=self.label)
        charts = []
        ws.column_dimensions['A'].best_fit = True
        ws.append(['Klass:', self.klass])
        ws.append(['Prov:', self.label])
        if self.comment:
            ws.append([self.comment])
        else:
            ws.append(['Inget avvikande noterat vid analys!'])
        ws.append([])  # 1 blank rows

        ws.append(['Absorbtionsmätning'])
        ws.cell(ws.max_row, 1).font = ft_heading
        write_spectrum(self, 'abs')
        ws.append([])  # 1 blank rows

        ws.append(['Flourescensmätning med 350 nm excitationsvåglängd'])
        ws.cell(ws.max_row, 1).font = ft_heading
        write_spectrum(self, 'ex350')
        ws.append([])  # 1 blank rows

        ws.append(['Flourescensmätning med 400 nm excitationsvåglängd'])
        ws.cell(ws.max_row, 1).font = ft_heading
        write_spectrum(self, 'ex400')

        anchor_row = ws.max_row+2
        anchor_cols = iter(['A', 'J', 'T'])
        for chart in charts:
            ws.add_chart(chart, '{}{}'.format(next(anchor_cols), anchor_row))


class CQDSpectrum:
    """Clas describing a series of one or more spectra from a Carbon Quantum Dot sample"""
    def __init__(self, wl_vector, y_vectors, meta_data):
        self.wl_vector = wl_vector  # numpy array of wavelength values
        self.y_vectors = y_vectors  # dictionary of numpy arrays containing y-value (absorbance or fluorescence)
        self.meta_data = meta_data  # dictionary containing metadata of spectra

    @classmethod
    def spectrum_from_xl_data(cls, start_t, end_t, attrib_col, value_col, rows):
        """
        Creates a CQDSpectrum object from data extracted from an excel file
        :param start_t: String: start time of measurement
        :param end_t: String: end time of measurement
        :param attrib_col: List: attributes of measurement
        :param value_col: List: values corresponding to attributes of measurement
        :param rows: List[List]: rows of spectral data
        :return: CQDSpectrum object
        """
        mode = value_col[attrib_col.index('Mode')]
        if mode == 'Absorbance':
            mode = 'Abs'
        elif mode == 'Fluorescence Top Reading':
            mode = 'Flu'
        else:
            print('{} is not a valid mode'.format(mode))
            return None

        meta_data = {'start_t': time.strptime(start_t, '%Y-%m-%d %H:%M:%S'),
                     'end_t': time.strptime(end_t, '%Y-%m-%d %H:%M:%S')}
        wl_vector = row_to_np_array(rows[0])
        y_vectors = {}
        if mode == 'Abs':
            y_vectors = {'Abs': row_to_np_array(rows[1])}
        elif mode == 'Flu':
            meta_data['gain'] = value_col[attrib_col.index('Gain')]
            meta_data['ex_wl'] = value_col[attrib_col.index('Excitation Wavelength')]
            y_vectors = {'Aq': row_to_np_array(rows[1]),
                         'Cu': row_to_np_array(rows[2]),
                         'Fe': row_to_np_array(rows[3]),
                         'Cd': row_to_np_array(rows[4])}

        return CQDSpectrum(wl_vector, y_vectors, meta_data)

